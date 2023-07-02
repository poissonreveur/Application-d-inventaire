import tkinter as tk
from tkinter import messagebox, StringVar
from openpyxl import load_workbook
from datetime import datetime
from PIL import ImageTk, Image

def ajouter_outil():
    nom_outil = entry_nom.get()
    date_emprunt = entry_emprunt.get()
    date_depot = entry_depot.get()
    etat_materiel = var_etat.get()
    emprunteur = entry_emprunteur.get()
        
    if not nom_outil:
        messagebox.showerror("Erreur", "Veuillez renseigner le nom de l'outil.")
        return
    
    # Formater les dates
    try:
        if date_emprunt:
            #feuille.cell(row=prochaine_ligne, column=2, value=datetime.strptime(date_emprunt, "%d/%m/%Y").strftime("%d/%m/%Y"))
            date_emprunt = datetime.strptime(date_emprunt, "%d/%m/%Y").strftime("%d/%m/%Y")
        date_depot = datetime.strptime(date_depot, "%d/%m/%Y").strftime("%d/%m/%Y")
    except ValueError:
        messagebox.showerror("Erreur", "Veuillez entrer les dates au format dd/mm/yyyy.")
        return

    # Charger le fichier Excel existant
    try:
        classeur = load_workbook("fichier.xlsx")
        feuille = classeur.active
    except:
        messagebox.showerror("Erreur", "Erreur lors du chargement du fichier Excel.")
        return

    # Trouver la première ligne vide
    prochaine_ligne = feuille.max_row + 1

    # Copier les données existantes à partir du fichier Excel
    for colonne in range(1, 6):
        cellule_source = feuille.cell(row=1, column=colonne)
        cellule_destination = feuille.cell(row=prochaine_ligne, column=colonne)
        cellule_destination.value = cellule_source.value

    # Ajouter les nouvelles informations dans la ligne suivante
    feuille.cell(row=prochaine_ligne, column=1, value=nom_outil)
    feuille.cell(row=prochaine_ligne, column=2, value=date_emprunt)
    feuille.cell(row=prochaine_ligne, column=3, value=date_depot)
    feuille.cell(row=prochaine_ligne, column=4, value=etat_materiel)
    feuille.cell(row=prochaine_ligne, column=5, value=emprunteur)

    # Sauvegarder le fichier Excel
    classeur.save("fichier.xlsx")

    # Effacer les champs après avoir ajouté les informations
    entry_nom.delete(0, tk.END)
    entry_emprunt.delete(0, tk.END)
    entry_depot.delete(0, tk.END)
    entry_emprunteur.delete(0, tk.END)

def rechercher():
    recherche = entry_recherche.get().lower()

    # Charger le fichier Excel existant
    try:
        classeur = load_workbook("fichier.xlsx")
        feuille = classeur.active
    except:
        messagebox.showerror("Erreur", "Erreur lors du chargement du fichier Excel.")
        return

    # Réinitialiser la zone d'affichage
    text_affichage.delete(1.0, tk.END)

    # Vérifier si l'outil est présent dans le fichier
    outil_present = False

    # Parcourir les données et afficher les correspondances
    for ligne in range(2, feuille.max_row + 1):
        nom_outil = feuille.cell(row=ligne, column=1).value
        if nom_outil is not None and recherche in nom_outil.lower():
            outil_present = True
            date_emprunt = feuille.cell(row=ligne, column=2).value
            date_depot = feuille.cell(row=ligne, column=3).value
            etat_materiel = feuille.cell(row=ligne, column=4).value
            emprunteur = feuille.cell(row=ligne, column=5).value

            # Afficher les informations correspondantes
            texte = f"Nom de l'outil: {nom_outil}\n"
            texte += f"Date d'emprunt: {date_emprunt}\n"
            texte += f"Date de dépôt: {date_depot}\n"
            texte += f"État du matériel: {etat_materiel}\n"
            texte += f"Emprunteur: {emprunteur}\n"
            texte += "------------------------\n"

            text_affichage.insert(tk.END, texte)

    # Afficher un message si l'outil n'est pas présent
    if not outil_present:
        messagebox.showinfo("Information", "L'outil n'est pas présent dans le fichier.")


def rechercher():
    recherche = entry_recherche.get().lower()

    # Charger le fichier Excel existant
    try:
        classeur = load_workbook("fichier.xlsx")
        feuille = classeur.active
    except:
        messagebox.showerror("Erreur", "Erreur lors du chargement du fichier Excel.")
        return

    # Réinitialiser la zone d'affichage
    text_affichage.delete(1.0, tk.END)

    # Vérifier si l'outil est présent dans le fichier
    outil_present = False

    # Parcourir les données et afficher les correspondances
    for ligne in range(2, feuille.max_row + 1):
        nom_outil = feuille.cell(row=ligne, column=1).value
        if nom_outil is not None and recherche in nom_outil.lower():
            outil_present = True
            date_emprunt = feuille.cell(row=ligne, column=2).value
            date_depot = feuille.cell(row=ligne, column=3).value
            etat_materiel = feuille.cell(row=ligne, column=4).value
            emprunteur = feuille.cell(row=ligne, column=5).value

            # Afficher les informations correspondantes
            texte = f"Nom de l'outil: {nom_outil}\n"
            texte += f"Date d'emprunt: {date_emprunt}\n"
            texte += f"Date de dépôt: {date_depot}\n"
            texte += f"État du matériel: {etat_materiel}\n"
            texte += f"Emprunteur: {emprunteur}\n"
            texte += "------------------------\n"

            text_affichage.insert(tk.END, texte)

    # Afficher un message si l'outil n'est pas présent
    if not outil_present:
        messagebox.showinfo("Information", "L'outil n'est pas présent dans le fichier.")


def ouvrir_fenetre_modifier():
    # Création de la fenêtre de modification
    fenetre_modifier = tk.Toplevel(fenetre)
    fenetre_modifier.title("Modifier un outil")
    fenetre_modifier.configure(background='#2D2D36')
    fenetre_modifier.iconbitmap('icone.ico')
    fenetre_modifier.geometry("300x300")


    def obtenir_informations():
        # Charger le fichier Excel existant
        try:
            classeur = load_workbook("fichier.xlsx")
            feuille = classeur.active
        except:
            messagebox.showerror("Erreur", "Erreur lors du chargement du fichier Excel.")
            fenetre_modifier.destroy()
            return

        # Obtenir la liste des outils disponibles
        outils_disponibles = []
        for ligne in range(2, feuille.max_row + 1):
            nom_outil = feuille.cell(row=ligne, column=1).value
            if nom_outil is not None:
                outils_disponibles.append(nom_outil)

        return outils_disponibles

    def valider_modification():
        # Récupérer les valeurs saisies dans les champs de saisie
        outil_modifie = outils.get()
        emprunteur_modifie = entry_emprunteur.get()
        date_emprunt_modifie = entry_date_emprunt.get()
        date_depot_modifie = entry_date_depot.get()
        etat_materiel_modifie = var_etat.get()

        # Vérifier si les dates sont valides
        try:
            if date_emprunt_modifie:
                date_emprunt_modifie = datetime.strptime(date_emprunt_modifie, "%d/%m/%Y").strftime("%d/%m/%Y")
            date_depot_modifie = datetime.strptime(date_depot_modifie, "%d/%m/%Y").strftime("%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Erreur", "Veuillez entrer les dates au format dd/mm/yyyy.")
            return

        # Charger le fichier Excel existant
        try:
            classeur = load_workbook("fichier.xlsx")
            feuille = classeur.active
        except:
            messagebox.showerror("Erreur", "Erreur lors du chargement du fichier Excel.")
            return

        # Parcourir les données et mettre à jour les informations de l'outil sélectionné
        for ligne in range(2, feuille.max_row + 1):
            nom_outil = feuille.cell(row=ligne, column=1).value
            if nom_outil == outil_modifie:
                feuille.cell(row=ligne, column=2, value=date_emprunt_modifie)
                feuille.cell(row=ligne, column=3, value=date_depot_modifie)
                feuille.cell(row=ligne, column=4, value=etat_materiel_modifie)
                feuille.cell(row=ligne, column=5, value=emprunteur_modifie)

                # Sauvegarder le fichier Excel
                classeur.save("fichier.xlsx")

                messagebox.showinfo("Information", "Les modifications ont été enregistrées avec succès.")
                fenetre_modifier.destroy()
                return

        messagebox.showerror("Erreur", "L'outil sélectionné n'a pas été trouvé dans le fichier.")
        fenetre_modifier.destroy()
    # Obtenir la liste des outils disponibles
    outils_disponibles = obtenir_informations()

    if not outils_disponibles:
        messagebox.showinfo("Information", "Aucun outil disponible pour la modification.")
        return

    # Création des étiquettes et des champs de saisie pour la modification
    label_outil = tk.Label(fenetre_modifier, text="Outil à modifier:", background='#2D2D36', fg='white')
    label_outil.pack()

    # Création de la liste déroulante pour sélectionner l'outil à modifier
    outils = tk.StringVar(fenetre_modifier)
    outils.set(outils_disponibles[0])  # Valeur par défaut

    menu_outils = tk.OptionMenu(fenetre_modifier, outils, *outils_disponibles)
    menu_outils.config(background='#2D2D36', fg='white', highlightthickness=0)
    menu_outils.pack()

    label_emprunteur = tk.Label(fenetre_modifier, text="Emprunteur:", background='#2D2D36', fg='white')
    label_emprunteur.pack()

    entry_emprunteur = tk.Entry(fenetre_modifier)
    entry_emprunteur.pack()

    label_date_emprunt = tk.Label(fenetre_modifier, text="Date d'emprunt:", background='#2D2D36', fg='white')
    label_date_emprunt.pack()

    entry_date_emprunt = tk.Entry(fenetre_modifier)
    entry_date_emprunt.pack()

    label_date_depot = tk.Label(fenetre_modifier, text="Date de dépôt:", background='#2D2D36', fg='white')
    label_date_depot.pack()

    entry_date_depot = tk.Entry(fenetre_modifier)
    entry_date_depot.pack()

    # Création du menu déroulant pour l'état du matériel
    label_etat_materiel = tk.Label(fenetre_modifier, text="État du matériel:", background='#2D2D36', fg='white')
    label_etat_materiel.pack()

    var_etat = tk.StringVar(fenetre_modifier)
    var_etat.set("Bon état")  # Valeur par défaut

    menu_etat = tk.OptionMenu(fenetre_modifier, var_etat, "Bon état", "Endommagé", "À réparer")
    menu_etat.config(background='#2D2D36', fg='white', highlightthickness=0)
    menu_etat.pack()

    # Boutons de validation et d'annulation
    bouton_valider = tk.Button(fenetre_modifier, text="Valider", command=valider_modification, background='#E04750', fg='white')
    bouton_valider.pack()

    bouton_annuler = tk.Button(fenetre_modifier, text="Annuler", command=fenetre_modifier.destroy, background='#E04750', fg='white')
    bouton_annuler.pack()

def valider_champs():
    nom_outil = entry_nom.get()
    date_depot = entry_depot.get()

    if nom_outil and date_depot:
        bouton_ajouter.configure(state=tk.NORMAL)
    else:
        bouton_ajouter.configure(state=tk.DISABLED)


# Création de la fenêtre principale
fenetre = tk.Tk()
fenetre.title("Recensement des Outils")
fenetre.iconbitmap('icone.ico')
fenetre.configure(background='#2D2D36')

# Création des étiquettes et des champs de saisie
label_nom = tk.Label(fenetre, text="Nom de l'outil:", background='#2D2D36', fg='white')
entry_nom = tk.Entry(fenetre)
entry_nom.bind("<<Modified>>", lambda event: valider_champs())

label_emprunt = tk.Label(fenetre, text="Date d'emprunt:", background='#2D2D36', fg='white')
entry_emprunt = tk.Entry(fenetre)

label_depot = tk.Label(fenetre, text="Date de dépôt:", background='#2D2D36', fg='white')
entry_depot = tk.Entry(fenetre)
entry_depot.bind("<<Modified>>", lambda event: valider_champs())

label_emprunteur = tk.Label(fenetre, text="Emprunteur:", background='#2D2D36', fg='white')
entry_emprunteur = tk.Entry(fenetre)

# Création de la barre de recherche
label_recherche = tk.Label(fenetre, text="Recherche :", background='#2D2D36', fg='white')
label_recherche.pack()

entry_recherche = tk.Entry(fenetre)
entry_recherche.pack()

bouton_rechercher = tk.Button(fenetre, text="Rechercher", command=rechercher, background='#E04750', fg='white')
bouton_rechercher.pack()

# Création de la zone d'affichage des résultats
text_affichage = tk.Text(fenetre, background='white')
text_affichage.pack()

# Bouton pour ajouter l'outil
bouton_ajouter = tk.Button(fenetre, text="Ajouter l'outil", command=ajouter_outil, background='#E04750', fg='white')

# Placement des éléments dans la fenêtre
label_nom.pack()
entry_nom.pack()
label_emprunt.pack()
entry_emprunt.pack()
label_depot.pack()
entry_depot.pack()
label_emprunteur.pack()
entry_emprunteur.pack()

# Création du menu déroulant pour l'état du matériel
var_etat = StringVar(fenetre)
var_etat.set("Bon état")  # Valeur par défaut

label_etat_materiel = tk.Label(fenetre, text="État du matériel:", background='#2D2D36', fg='white')
label_etat_materiel.pack()

menu_etat = tk.OptionMenu(fenetre, var_etat, "Bon état", "Endommagé", "À réparer")
menu_etat.config(background='#2D2D36', fg='white', highlightthickness=0)
menu_etat.pack()

bouton_ajouter.pack(side=tk.BOTTOM)
bouton_ajouter = tk.Button(fenetre, text="Ajouter l'outil", command=ajouter_outil, background='#E04750', fg='white')
bouton_ajouter.configure(state=tk.DISABLED)

# Bouton pour ouvrir la fenêtre de modification
bouton_modifier = tk.Button(fenetre, text="Modifier un outil", command=ouvrir_fenetre_modifier, background='#E04750', fg='white')
bouton_modifier.pack(side=tk.BOTTOM)

# Boucle principale de l'application
fenetre.mainloop()
